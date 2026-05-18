"""
Build an S&P 500 value-opportunity spreadsheet.

This script:
1. Downloads the current S&P 500 company list from Wikipedia.
2. Pulls quarterly revenue/EPS, valuation metrics, and historical prices from
   Yahoo Finance through yfinance.
3. Includes same-quarter previous-year revenue/EPS for direct comparison.
4. Calculates year-over-year and two-year average growth metrics.
5. Saves all rows plus a sorted "Value Opportunities" view to Excel.

Install dependencies:
    pip install yfinance pandas wikipedia openpyxl lxml

The user requested yfinance, pandas, and wikipedia. openpyxl is included above
because pandas needs an Excel writer engine to create .xlsx files. lxml is
included because pandas uses it to parse Wikipedia's HTML table.
"""

from __future__ import annotations

import time
from datetime import datetime, timedelta
from io import StringIO
from typing import Optional

import pandas as pd
import wikipedia
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT_FILE = "sp500_value_opportunities.xlsx"
WIKIPEDIA_PAGE_TITLE = "List of S&P 500 companies"

# Yahoo Finance can throttle aggressive scraping. Increase this if you see
# intermittent failures or HTTP 429 responses.
REQUEST_DELAY_SECONDS = 1.25


OUTPUT_COLUMNS = [
    "Ticker",
    "Company Name",
    "Current Price",
    "Market Cap",
    "Revenue (latest quarter)",
    "Revenue (same quarter previous year)",
    "Revenue YoY Growth %",
    "EPS (latest quarter)",
    "EPS (same quarter previous year)",
    "EPS YoY Growth %",
    "Revenue Growth 2yr Avg %",
    "EPS Growth 2yr Avg %",
    "P/E Ratio",
    "PEG Ratio",
    "Price 1 Year Ago",
    "1 Year Price Change %",
    "Price 2 Years Ago",
    "2 Year Price Change %",
    "Price 3 Years Ago",
    "3 Year Price Change %",
]


def fetch_sp500_companies() -> pd.DataFrame:
    """Return the current S&P 500 ticker and company-name table from Wikipedia."""
    print("Downloading current S&P 500 company list from Wikipedia...")

    # The wikipedia package fetches the page HTML; pandas then parses the table.
    page = wikipedia.page(WIKIPEDIA_PAGE_TITLE, auto_suggest=False)
    tables = pd.read_html(StringIO(page.html()))

    # The first table on this page is the current constituents table.
    constituents = tables[0]
    constituents = constituents.rename(
        columns={
            "Symbol": "Ticker",
            "Security": "Company Name",
        }
    )

    return constituents[["Ticker", "Company Name"]].copy()


def to_yahoo_ticker(wikipedia_ticker: str) -> str:
    """Convert Wikipedia share-class tickers to Yahoo Finance format."""
    return wikipedia_ticker.replace(".", "-")


def safe_float(value) -> Optional[float]:
    """Convert numeric-ish values to float, returning None for missing values."""
    if value is None:
        return None

    try:
        if pd.isna(value):
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def safe_get(mapping, key: str):
    """Read a key from dict-like yfinance objects without assuming exact type."""
    if mapping is None:
        return None

    try:
        return mapping.get(key)
    except AttributeError:
        pass

    try:
        return mapping[key]
    except (KeyError, TypeError):
        return None


def percent_change(current: Optional[float], previous: Optional[float]) -> Optional[float]:
    """Return percentage change from previous to current, handling bad inputs."""
    current = safe_float(current)
    previous = safe_float(previous)

    if current is None or previous in (None, 0):
        return None

    return ((current - previous) / abs(previous)) * 100


def find_statement_row(statement: pd.DataFrame, possible_names: list[str]) -> Optional[str]:
    """Find the first matching row name in a yfinance financial statement."""
    if statement is None or statement.empty:
        return None

    normalized_index = {str(index).strip().lower(): index for index in statement.index}
    for name in possible_names:
        match = normalized_index.get(name.strip().lower())
        if match is not None:
            return match

    return None


def latest_value(statement: pd.DataFrame, row_name: Optional[str]) -> Optional[float]:
    """Return the latest reported value for a financial-statement row."""
    if row_name is None or statement is None or statement.empty:
        return None

    values = statement.loc[row_name].dropna()
    if values.empty:
        return None

    return safe_float(values.iloc[0])


def statement_value_at_offset(
    statement: pd.DataFrame,
    row_name: Optional[str],
    quarter_offset: int,
) -> Optional[float]:
    """Return a row value at a specific quarterly offset from the latest quarter."""
    if row_name is None or statement is None or statement.empty:
        return None

    values = statement.loc[row_name].dropna()
    if len(values) <= quarter_offset:
        return None

    return safe_float(values.iloc[quarter_offset])


def yoy_growth(statement: pd.DataFrame, row_name: Optional[str], quarter_offset: int = 0) -> Optional[float]:
    """
    Return year-over-year growth for a quarterly statement row.

    quarter_offset=0 means latest quarter versus the same quarter one year prior.
    quarter_offset=1 means previous quarter versus its same quarter one year prior.
    """
    if row_name is None or statement is None or statement.empty:
        return None

    values = statement.loc[row_name].dropna()
    needed_columns = quarter_offset + 5
    if len(values) < needed_columns:
        return None

    current = values.iloc[quarter_offset]
    prior_year = values.iloc[quarter_offset + 4]
    return percent_change(current, prior_year)


def two_year_average_yoy_growth(statement: pd.DataFrame, row_name: Optional[str]) -> Optional[float]:
    """
    Return the average of the latest two available quarterly YoY growth rates.

    yfinance commonly exposes only five quarters of quarterly statement data,
    which is enough for the latest YoY growth calculation but not always enough
    for a true eight-quarter CAGR. This helper uses the latest two YoY readings
    when available and falls back to the latest YoY reading when only five
    quarters are available.
    """
    growth_rates = [
        yoy_growth(statement, row_name, quarter_offset=0),
        yoy_growth(statement, row_name, quarter_offset=1),
    ]
    growth_rates = [rate for rate in growth_rates if rate is not None]

    if not growth_rates:
        return None

    return sum(growth_rates) / len(growth_rates)


def get_price_on_or_before(
    price_history: pd.DataFrame,
    target_date: datetime,
    lookback_days: int = 10,
) -> Optional[float]:
    """
    Return adjusted close on the nearest trading day on or before target_date.

    This handles weekends and market holidays by searching backward up to
    lookback_days. If yfinance did not provide an auto-adjusted Close column,
    the function also supports the older "Adj Close" column.
    """
    if price_history is None or price_history.empty:
        return None

    price_column = "Close" if "Close" in price_history.columns else "Adj Close"
    if price_column not in price_history.columns:
        return None

    index = price_history.index
    if getattr(index, "tz", None) is not None:
        # Keep date comparisons simple by removing timezone information.
        price_history = price_history.copy()
        price_history.index = price_history.index.tz_localize(None)

    window_start = target_date - timedelta(days=lookback_days)
    window = price_history.loc[
        (price_history.index.date >= window_start.date())
        & (price_history.index.date <= target_date.date())
    ]

    if window.empty:
        return None

    return safe_float(window[price_column].dropna().iloc[-1])


def get_company_metrics(ticker: str, company_name: str) -> dict:
    """Fetch and calculate all requested metrics for one S&P 500 company."""
    yahoo_ticker = to_yahoo_ticker(ticker)
    stock = yf.Ticker(yahoo_ticker)

    # fast_info is lighter than info for price/market cap, but not all fields
    # are always present. info remains useful for valuation ratios.
    fast_info = {}
    info = {}

    try:
        fast_info = stock.fast_info or {}
    except Exception:
        fast_info = {}

    try:
        info = stock.info or {}
    except Exception:
        info = {}

    current_price = safe_float(
        safe_get(fast_info, "lastPrice")
        or safe_get(info, "currentPrice")
        or safe_get(info, "regularMarketPrice")
        or safe_get(info, "previousClose")
    )
    market_cap = safe_float(safe_get(fast_info, "marketCap") or safe_get(info, "marketCap"))
    trailing_pe = safe_float(safe_get(info, "trailingPE"))
    peg_ratio = safe_float(safe_get(info, "pegRatio") or safe_get(info, "trailingPegRatio"))

    # yfinance exposes the same data through a few aliases. quarterly_financials
    # is retained here because it matches the user's request wording.
    quarterly_financials = stock.quarterly_financials
    quarterly_income_stmt = stock.quarterly_income_stmt
    statement = quarterly_income_stmt if not quarterly_income_stmt.empty else quarterly_financials

    revenue_row = find_statement_row(
        statement,
        [
            "Total Revenue",
            "Operating Revenue",
        ],
    )
    eps_row = find_statement_row(
        statement,
        [
            "Diluted EPS",
            "Basic EPS",
            "Normalized Diluted EPS",
        ],
    )

    latest_revenue = latest_value(statement, revenue_row)
    previous_year_revenue = statement_value_at_offset(statement, revenue_row, quarter_offset=4)
    latest_eps = latest_value(statement, eps_row)
    previous_year_eps = statement_value_at_offset(statement, eps_row, quarter_offset=4)
    revenue_yoy = yoy_growth(statement, revenue_row)
    eps_yoy = yoy_growth(statement, eps_row)
    revenue_two_year_avg = two_year_average_yoy_growth(statement, revenue_row)
    eps_two_year_avg = two_year_average_yoy_growth(statement, eps_row)

    today = pd.Timestamp(datetime.now().date())
    target_1y = today - pd.DateOffset(years=1)
    target_2y = today - pd.DateOffset(years=2)
    target_3y = today - pd.DateOffset(years=3)
    start_date = target_3y - pd.DateOffset(days=30)

    price_history = stock.history(
        start=start_date.strftime("%Y-%m-%d"),
        end=(today + timedelta(days=1)).strftime("%Y-%m-%d"),
        auto_adjust=True,
    )

    price_1y = get_price_on_or_before(price_history, target_1y.to_pydatetime())
    price_2y = get_price_on_or_before(price_history, target_2y.to_pydatetime())
    price_3y = get_price_on_or_before(price_history, target_3y.to_pydatetime())

    return {
        "Ticker": ticker,
        "Company Name": company_name,
        "Current Price": current_price,
        "Market Cap": market_cap,
        "Revenue (latest quarter)": latest_revenue,
        "Revenue (same quarter previous year)": previous_year_revenue,
        "Revenue YoY Growth %": revenue_yoy,
        "EPS (latest quarter)": latest_eps,
        "EPS (same quarter previous year)": previous_year_eps,
        "EPS YoY Growth %": eps_yoy,
        "Revenue Growth 2yr Avg %": revenue_two_year_avg,
        "EPS Growth 2yr Avg %": eps_two_year_avg,
        "P/E Ratio": trailing_pe,
        "PEG Ratio": peg_ratio,
        "Price 1 Year Ago": price_1y,
        "1 Year Price Change %": percent_change(current_price, price_1y),
        "Price 2 Years Ago": price_2y,
        "2 Year Price Change %": percent_change(current_price, price_2y),
        "Price 3 Years Ago": price_3y,
        "3 Year Price Change %": percent_change(current_price, price_3y),
    }


def build_value_score(df: pd.DataFrame) -> pd.Series:
    """
    Score value opportunities: higher growth and lower price appreciation rank first.

    This is intentionally simple and transparent, not investment advice. The
    score rewards revenue/EPS growth and penalizes recent price appreciation,
    P/E, and PEG when available.
    """
    revenue_growth = df["Revenue Growth 2yr Avg %"].fillna(0)
    eps_growth = df["EPS Growth 2yr Avg %"].fillna(0)
    one_year_price_change = df["1 Year Price Change %"].fillna(0)
    pe_ratio = df["P/E Ratio"].fillna(df["P/E Ratio"].median())
    peg_ratio = df["PEG Ratio"].fillna(df["PEG Ratio"].median())

    # Clamp extreme ratios so one bad/missing feed value does not dominate.
    pe_penalty = pe_ratio.clip(lower=0, upper=100) * 0.15
    peg_penalty = peg_ratio.clip(lower=0, upper=5) * 3

    return (
        revenue_growth * 0.45
        + eps_growth * 0.45
        - one_year_price_change * 0.35
        - pe_penalty
        - peg_penalty
    )


def write_excel_file(df: pd.DataFrame, output_file: str) -> None:
    """Write all stocks and sorted value opportunities to a two-sheet Excel file."""
    all_stocks = df[OUTPUT_COLUMNS].copy()
    all_stocks = all_stocks.sort_values("Market Cap", ascending=False, na_position="last")

    opportunities = all_stocks.copy()
    opportunities["_Value Score"] = build_value_score(opportunities)
    opportunities = opportunities.sort_values("_Value Score", ascending=False)
    opportunities = opportunities.drop(columns=["_Value Score"])

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        all_stocks.to_excel(writer, sheet_name="All Stocks", index=False)
        opportunities.to_excel(writer, sheet_name="Value Opportunities", index=False)

    format_excel_file(output_file)


def format_excel_file(output_file: str) -> None:
    """Apply clean Excel display formatting to the generated workbook."""
    currency_cols = {
        "Current Price",
        "Revenue (latest quarter)",
        "Revenue (same quarter previous year)",
        "EPS (latest quarter)",
        "EPS (same quarter previous year)",
        "Price 1 Year Ago",
        "Price 2 Years Ago",
        "Price 3 Years Ago",
    }
    large_currency_cols = {"Market Cap"}
    percentage_cols = {
        "Revenue YoY Growth %",
        "EPS YoY Growth %",
        "Revenue Growth 2yr Avg %",
        "EPS Growth 2yr Avg %",
        "1 Year Price Change %",
        "2 Year Price Change %",
        "3 Year Price Change %",
    }
    ratio_cols = {"P/E Ratio", "PEG Ratio"}

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D1D5DB")
    border = Border(bottom=thin_gray)

    workbook = load_workbook(output_file)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.sheet_view.showGridLines = False

        headers = [cell.value for cell in worksheet[1]]

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)

            if header == "Ticker":
                worksheet.column_dimensions[col_letter].width = 12
            elif header == "Company Name":
                worksheet.column_dimensions[col_letter].width = 34
            elif header in large_currency_cols:
                worksheet.column_dimensions[col_letter].width = 18
            elif header in currency_cols or header in percentage_cols:
                worksheet.column_dimensions[col_letter].width = 18
            elif header in ratio_cols:
                worksheet.column_dimensions[col_letter].width = 12
            else:
                worksheet.column_dimensions[col_letter].width = 16

            for row_idx in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(vertical="center")

                if header in large_currency_cols:
                    cell.number_format = "$#,##0"
                elif header in currency_cols:
                    cell.number_format = "$#,##0.00"
                elif header in percentage_cols:
                    # Values are stored as percentage points: 12.3 displays as 12.30%.
                    cell.number_format = r"0.00\%"
                elif header in ratio_cols:
                    cell.number_format = "0.00"

        for row_idx in range(1, worksheet.max_row + 1):
            worksheet.row_dimensions[row_idx].height = 20
        worksheet.row_dimensions[1].height = 36

    workbook.save(output_file)


def main() -> None:
    """Run the full S&P 500 data collection process."""
    companies = fetch_sp500_companies()
    total = len(companies)
    rows = []

    print(f"Found {total} S&P 500 constituents.")

    for index, row in companies.iterrows():
        ticker = row["Ticker"]
        company_name = row["Company Name"]
        try:
            metrics = get_company_metrics(ticker, company_name)
            rows.append(metrics)
            print(f"Processed {index + 1}/{total} tickers: {ticker} - {company_name}")
        except Exception as exc:
            # One bad ticker should not stop the entire run. Keep the message
            # short so long S&P 500 runs remain readable.
            print(f"Skipped {index + 1}/{total} tickers: {ticker} - {exc}")
        finally:
            time.sleep(REQUEST_DELAY_SECONDS)

    if not rows:
        raise RuntimeError("No ticker data was collected; Excel file was not created.")

    df = pd.DataFrame(rows)
    df = df.reindex(columns=OUTPUT_COLUMNS)
    write_excel_file(df, OUTPUT_FILE)

    print(f"Saved {len(df)} rows to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
